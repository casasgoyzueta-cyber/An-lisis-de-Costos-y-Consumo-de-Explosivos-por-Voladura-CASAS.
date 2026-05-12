import io
import math
import warnings
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

warnings.filterwarnings("ignore")

plt.rcParams.update({
    "figure.dpi": 130,
    "figure.facecolor": "#F8F9FA",
    "axes.facecolor": "#FFFFFF",
    "axes.grid": True,
    "grid.alpha": 0.35,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "font.family": "DejaVu Sans",
    "font.size": 10,
})

# ─────────────────────────────────────────────
# PARÁMETROS
# ─────────────────────────────────────────────
PARAM_INFO = {
    "rho_r": {"label": "Densidad de la roca in situ", "symbol": "ρr", "unit": "g/cm³",
               "default": 2.70, "std_frac": 0.07, "stochastic": True,
               "hint": "Muy blanda 1.8–2.2 · Media 2.2–2.5 · Dura 2.5–2.9 · Ultra-dura 3.0–4.0"},
    "rho_e": {"label": "Densidad del explosivo", "symbol": "ρe", "unit": "g/cm³",
               "default": 1.10, "std_frac": 0.04, "stochastic": True,
               "hint": "ANFO seco 0.80 · ANFO pesado 1.00–1.10 · Emulsión bombeada 1.15–1.25"},
    "e_expl": {"label": "Energía específica del explosivo", "symbol": "e_expl", "unit": "MJ/kg",
                "default": 3.50, "std_frac": 0.05, "stochastic": True,
                "hint": "ANFO seco 3.85 · ANFO pesado 3.50 · Emulsión 3.20–3.30 · Rango 2.5–4.2"},
    "Dh": {"label": "Diámetro del taladro", "symbol": "Dh", "unit": "mm",
            "default": 165.0, "std_frac": 0.03, "stochastic": True,
            "hint": "Jack-leg 89–115 · Jumbo mediano 115–165 · Rotary grande 251–381"},
    "H": {"label": "Altura del banco", "symbol": "H", "unit": "m",
           "default": 10.0, "std_frac": 0.0, "stochastic": False,
           "hint": "Subterránea 3–6 · Banco bajo 5–10 · Estándar 10–15 · Alto 15–20"},
    "alpha": {"label": "Ángulo del taladro (respecto a la horizontal)", "symbol": "α", "unit": "°",
               "default": 80.0, "std_frac": 0.0, "stochastic": False,
               "hint": "Vertical 90° · Inclinación leve 75–85° · Moderada 65–75° · Pronunciada 55–65°"},
    "Ks": {"label": "Coeficiente de espaciamiento", "symbol": "Ks", "unit": "adim.",
            "default": 1.15, "std_frac": 0.0, "stochastic": False,
            "hint": "Roca fracturada 1.0–1.1 · Estándar Ash 1963: 1.15 · Competente 1.2–1.5"},
    "Kj": {"label": "Coeficiente de sobreperforación", "symbol": "Kj", "unit": "adim.",
            "default": 0.30, "std_frac": 0.0, "stochastic": False,
            "hint": "Roca blanda 0.20–0.25 · Estándar Langefors 0.30 · Roca dura 0.35–0.40"},
    "Kt": {"label": "Coeficiente de taco (stemming)", "symbol": "Kt", "unit": "adim.",
            "default": 0.75, "std_frac": 0.0, "stochastic": False,
            "hint": "Máx. eficiencia 0.70–0.75 · Estándar 0.75–0.80 · Control flyrock 0.90–1.00"},
    "Ce": {"label": "Costo unitario del explosivo principal", "symbol": "Ce", "unit": "USD/kg",
            "default": 0.95, "std_frac": 0.10, "stochastic": True,
            "hint": "ANFO seco 0.45–0.70 · ANFO pesado 0.80–1.20 · Emulsión 0.90–2.50"},
    "C_init": {"label": "Costo unitario del iniciador", "symbol": "C_init", "unit": "USD/u",
                "default": 4.50, "std_frac": 0.08, "stochastic": True,
                "hint": "Fulminante 0.50–1.50 · Nonel 2.00–8.00 · Electrónico 8.00–25.00"},
    "C_CD": {"label": "Costo del cordón detonante", "symbol": "C_CD", "unit": "USD/m",
              "default": 0.20, "std_frac": 0.0, "stochastic": False,
              "hint": "Ligero 0.10–0.18 · Estándar 0.18–0.28 · Pesado 0.28–0.40 · Sin cordón 0.00"},
    "Cp": {"label": "Costo de perforación (todo incluido)", "symbol": "Cp", "unit": "USD/m",
            "default": 18.00, "std_frac": 0.10, "stochastic": True,
            "hint": "Jack-leg 6–15 · Jumbo mediano 12–22 · Rotary mediano 18–32 · Grande 25–45"},
    "Nt": {"label": "Número total de taladros por disparo/banco", "symbol": "Nt", "unit": "u",
            "default": 48, "std_frac": 0.0, "stochastic": False,
            "hint": "Disparo pequeño 10–30 · Mediano 30–80 · Grande 80–200 · Muy grande 200–600"},
    "N_init": {"label": "Número de iniciadores por taladro", "symbol": "N_init", "unit": "u",
                "default": 1, "std_frac": 0.0, "stochastic": False,
                "hint": "Simple 1 · Doble >15 m 2 · Triple roca muy dura 2–3"},
    "Fop": {"label": "Factor de costos operacionales adicionales (overhead)", "symbol": "Fop", "unit": "%",
             "default": 12.0, "std_frac": 0.0, "stochastic": False,
             "hint": "Mínimo 8–10 · Estándar 10–15 · Alto 15–20"},
    "n_sim": {"label": "Número de simulaciones Monte Carlo", "symbol": "N_sim", "unit": "u",
               "default": 10000, "std_frac": 0.0, "stochastic": False,
               "hint": "Rápido 5,000 · Estándar 10,000 · Alta precisión 50,000"},
}

# ─────────────────────────────────────────────
# FUNCIONES DE CÁLCULO
# ─────────────────────────────────────────────
def interpret_burden(B):
    if B < 1.0:
        return "PEQUEÑO", "Burden muy corto — riesgo de fragmentación excesiva y flyrock."
    elif B < 2.0:
        return "NORMAL-BAJO", "Burden dentro del rango para diámetros medianos."
    elif B < 3.5:
        return "NORMAL", "Burden adecuado para banco estándar con explosivo de densidad media."
    elif B < 5.0:
        return "GRANDE", "Burden grande — verificar que el explosivo tenga energía suficiente."
    else:
        return "MUY GRANDE", "Burden excesivo — revisar diámetro o tipo de explosivo."

def interpret_fc(FC):
    if FC < 0.10:
        return "MUY BAJO", "verde", "Consumo muy eficiente — roca blanda o explosivo muy potente."
    elif FC < 0.25:
        return "BAJO", "verde", "Roca blanda a media. Voladura económica."
    elif FC < 0.40:
        return "MODERADO", "amarillo", "Roca media a dura. Consumo estándar en cielo abierto."
    elif FC < 0.60:
        return "ALTO", "naranja", "Roca dura (granito/basalto). Considerar diámetro mayor."
    else:
        return "MUY ALTO", "rojo", "Roca ultra-dura. Optimizar malla o cambiar explosivo."

def interpret_cuv(CUV):
    if CUV < 1.0:
        return "EXCELENTE", "verde", "Costo muy bajo — alta eficiencia de voladura."
    elif CUV < 2.5:
        return "BUENO", "verde", "Costo competitivo — operación eficiente."
    elif CUV < 4.0:
        return "MODERADO", "amarillo", "Costo estándar para minas a cielo abierto medianas."
    elif CUV < 6.0:
        return "ALTO", "naranja", "Costo elevado — revisar malla y tipo de explosivo."
    else:
        return "MUY ALTO", "rojo", "Costo crítico — optimización urgente del diseño de voladura."

def interpret_rbs(RBS):
    if RBS < 90:
        return "INFERIOR AL ANFO", "El explosivo entrega menos energía por m³ que el ANFO."
    elif RBS <= 110:
        return "SIMILAR AL ANFO", "Energía volumétrica equivalente al ANFO de referencia."
    elif RBS <= 130:
        return "SUPERIOR AL ANFO", "Buena energía por metro de taladro."
    else:
        return "MUY SUPERIOR AL ANFO", "Explosivo muy energético por volumen."

def calc_geometry(p):
    B = 0.012 * math.sqrt(2.0 * p["rho_e"] / p["rho_r"]) * p["Dh"]
    S = p["Ks"] * B
    J = p["Kj"] * B
    Ltaco = p["Kt"] * B
    alpha_rad = math.radians(p["alpha"])
    Lt = p["H"] / math.sin(alpha_rad) + J
    Lc = max(Lt - Ltaco, 0.01)
    return {"B": B, "S": S, "J": J, "Ltaco": Ltaco, "Lt": Lt, "Lc": Lc}

def calc_explosive_mass(p, geom):
    r = p["Dh"] / 2000.0
    return math.pi * r**2 * geom["Lc"] * p["rho_e"] * 1000.0

def calc_load_factor(p, geom, Qe):
    Vtaladro = geom["B"] * geom["S"] * p["H"]
    Ttaladro = Vtaladro * p["rho_r"]
    FC = Qe / Ttaladro if Ttaladro > 0 else 0.0
    return {"Vtaladro": Vtaladro, "Ttaladro": Ttaladro, "FC": FC}

def calc_rws_rbs(p):
    e_ANFO, rho_ANFO = 3.85, 0.80
    RWS = (p["e_expl"] / e_ANFO) * 100.0
    RBS = RWS * (p["rho_e"] / rho_ANFO)
    B_equiv_factor = (RBS / 100.0) ** (1.0 / 3.0)
    return {"RWS": RWS, "RBS": RBS, "B_equiv_factor": B_equiv_factor}

def calc_costs_per_hole(p, geom, Qe):
    Ce_t = Qe * p["Ce"] + geom["Lc"] * p["C_CD"] + p["N_init"] * p["C_init"]
    Cp_t = geom["Lt"] * p["Cp"]
    return {"Ce_t": Ce_t, "Cp_t": Cp_t}

def calc_bank_totals(p, geom, costs, lf):
    CT = p["Nt"] * (costs["Ce_t"] + costs["Cp_t"]) * (1.0 + p["Fop"] / 100.0)
    Ttotal = p["Nt"] * lf["Ttaladro"]
    Vtotal = p["Nt"] * lf["Vtaladro"]
    CUV_t = CT / Ttotal if Ttotal > 0 else 0.0
    CUV_vol = CT / Vtotal if Vtotal > 0 else 0.0
    return {"CT": CT, "Ttotal": Ttotal, "Vtotal": Vtotal, "CUV_t": CUV_t, "CUV_vol": CUV_vol}

def calc_all(p):
    geom = calc_geometry(p)
    Qe = calc_explosive_mass(p, geom)
    lf = calc_load_factor(p, geom, Qe)
    lf["Qe"] = Qe
    energy = calc_rws_rbs(p)
    costs = calc_costs_per_hole(p, geom, Qe)
    bank = calc_bank_totals(p, geom, costs, lf)
    return {"geom": geom, "Qe": Qe, "lf": lf, "energy": energy, "costs": costs, "bank": bank}

def _sample_truncated_normal(mu, sigma, n, low_frac=0.5, high_frac=2.0):
    a = (mu * low_frac - mu) / sigma
    b = (mu * high_frac - mu) / sigma
    return stats.truncnorm.rvs(a, b, loc=mu, scale=sigma, size=n)

def monte_carlo(p):
    N = p["n_sim"]
    rng = np.random.default_rng(42)
    rho_r_s = _sample_truncated_normal(p["rho_r"], p["rho_r"] * PARAM_INFO["rho_r"]["std_frac"], N)
    rho_e_s = _sample_truncated_normal(p["rho_e"], p["rho_e"] * PARAM_INFO["rho_e"]["std_frac"], N)
    e_expl_s = _sample_truncated_normal(p["e_expl"], p["e_expl"] * PARAM_INFO["e_expl"]["std_frac"], N)
    Dh_s = rng.uniform(p["Dh"] - 5.0, p["Dh"] + 5.0, N)
    Ce_s = _sample_truncated_normal(p["Ce"], p["Ce"] * PARAM_INFO["Ce"]["std_frac"], N)
    C_init_s = _sample_truncated_normal(p["C_init"], p["C_init"] * PARAM_INFO["C_init"]["std_frac"], N)
    Cp_s = _sample_truncated_normal(p["Cp"], p["Cp"] * PARAM_INFO["Cp"]["std_frac"], N)

    H = p["H"]; alpha_rad = math.radians(p["alpha"])
    Ks = p["Ks"]; Kj = p["Kj"]; Kt = p["Kt"]
    Nt = p["Nt"]; N_init = p["N_init"]; C_CD = p["C_CD"]; Fop = p["Fop"]

    B_s = 0.012 * np.sqrt(2.0 * rho_e_s / rho_r_s) * Dh_s
    S_s = Ks * B_s; J_s = Kj * B_s; Ltaco_s = Kt * B_s
    Lt_s = H / math.sin(alpha_rad) + J_s
    Lc_s = np.maximum(Lt_s - Ltaco_s, 0.01)

    r_s = Dh_s / 2000.0
    Qe_s = np.pi * r_s**2 * Lc_s * rho_e_s * 1000.0

    Vtaladro_s = B_s * S_s * H
    Ttaladro_s = Vtaladro_s * rho_r_s
    FC_s = np.where(Ttaladro_s > 0, Qe_s / Ttaladro_s, 0.0)

    e_ANFO, rho_ANFO = 3.85, 0.80
    RWS_s = (e_expl_s / e_ANFO) * 100.0
    RBS_s = RWS_s * (rho_e_s / rho_ANFO)

    Ce_t_s = Qe_s * Ce_s + Lc_s * C_CD + N_init * C_init_s
    Cp_t_s = Lt_s * Cp_s
    CT_s = Nt * (Ce_t_s + Cp_t_s) * (1.0 + Fop / 100.0)
    Ttotal_s = Nt * Ttaladro_s
    CUV_s = np.where(Ttotal_s > 0, CT_s / Ttotal_s, 0.0)
    Vtotal_s = Nt * Vtaladro_s
    CUV_vol_s = np.where(Vtotal_s > 0, CT_s / Vtotal_s, 0.0)

    stochastic_inputs = {"rho_r": rho_r_s, "rho_e": rho_e_s, "e_expl": e_expl_s,
                          "Dh": Dh_s, "Ce": Ce_s, "C_init": C_init_s, "Cp": Cp_s}
    outputs = {"B": B_s, "Qe": Qe_s, "FC": FC_s, "RWS": RWS_s, "RBS": RBS_s,
               "CT": CT_s, "CUV_t": CUV_s, "CUV_vol": CUV_vol_s,
               "Ttotal": Ttotal_s, "Lt": Lt_s, "Lc": Lc_s}

    def stats_of(arr):
        return {"mean": float(np.mean(arr)), "std": float(np.std(arr)),
                "p5": float(np.percentile(arr, 5)), "p50": float(np.percentile(arr, 50)),
                "p90": float(np.percentile(arr, 90)), "p95": float(np.percentile(arr, 95))}

    mc_stats = {k: stats_of(v) for k, v in outputs.items()}
    correlations = {iname: float(stats.pearsonr(iarr, CUV_s)[0])
                    for iname, iarr in stochastic_inputs.items()}

    return {"stochastic_inputs": stochastic_inputs, "outputs": outputs,
            "mc_stats": mc_stats, "correlations": correlations, "n_sim": N}

# ─────────────────────────────────────────────
# GRÁFICOS
# ─────────────────────────────────────────────
def fig_histogram_cdf(arr, label, unit, det_value):
    fig, axes = plt.subplots(1, 2, figsize=(12, 4.5))
    fig.suptitle(f"Distribución MC: {label}", fontsize=11, fontweight="bold")
    n_bins = min(60, max(20, int(np.sqrt(len(arr)))))
    ax1 = axes[0]
    ax1.hist(arr, bins=n_bins, color="#2563EB", alpha=0.75, edgecolor="white", linewidth=0.4)
    ax1.axvline(det_value, color="#DC2626", lw=2, ls="--", label=f"Determinístico={det_value:.3f}")
    ax1.axvline(np.percentile(arr, 5), color="#F59E0B", lw=1.5, ls=":", label=f"P5={np.percentile(arr,5):.3f}")
    ax1.axvline(np.percentile(arr, 50), color="#10B981", lw=1.5, ls="-.", label=f"P50={np.percentile(arr,50):.3f}")
    ax1.axvline(np.percentile(arr, 95), color="#7C3AED", lw=1.5, ls=":", label=f"P95={np.percentile(arr,95):.3f}")
    ax1.set_xlabel(f"{label} [{unit}]"); ax1.set_ylabel("Frecuencia")
    ax1.set_title("Histograma"); ax1.legend(fontsize=8)
    ax2 = axes[1]
    sorted_arr = np.sort(arr)
    cdf = np.arange(1, len(sorted_arr) + 1) / len(sorted_arr)
    ax2.plot(sorted_arr, cdf * 100, color="#2563EB", lw=2)
    ax2.axvline(det_value, color="#DC2626", lw=2, ls="--", label=f"Det.={det_value:.3f}")
    ax2.axhline(50, color="#10B981", lw=1, ls="--", alpha=0.7)
    ax2.axhline(90, color="#F59E0B", lw=1, ls="--", alpha=0.7)
    ax2.set_xlabel(f"{label} [{unit}]"); ax2.set_ylabel("Probabilidad acumulada [%]")
    ax2.set_title("CDF Acumulada"); ax2.legend(fontsize=8)
    plt.tight_layout()
    return fig

def fig_tornado(correlations):
    labels_map = {"rho_r": "Densidad roca ρr", "rho_e": "Densidad explosivo ρe",
                  "e_expl": "Energía específica e_expl", "Dh": "Diámetro taladro Dh",
                  "Ce": "Costo explosivo Ce", "C_init": "Costo iniciador C_init",
                  "Cp": "Costo perforación Cp"}
    items = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)
    names = [labels_map.get(k, k) for k, _ in items]
    vals = [v for _, v in items]
    colors = ["#2563EB" if v >= 0 else "#DC2626" for v in vals]
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.suptitle("Diagrama de Tornado: Correlaciones de Pearson con CUV [USD/t]",
                 fontsize=10, fontweight="bold")
    bars = ax.barh(names, vals, color=colors, edgecolor="white", height=0.6)
    ax.axvline(0, color="black", lw=1)
    for bar, val in zip(bars, vals):
        ax.text(val + (0.01 if val >= 0 else -0.01), bar.get_y() + bar.get_height() / 2,
                f"{val:.3f}", va="center", ha="left" if val >= 0 else "right", fontsize=9)
    ax.set_xlabel("Coeficiente de correlación de Pearson r")
    pos_patch = mpatches.Patch(color="#2563EB", label="Positiva (↑ input → ↑ CUV)")
    neg_patch = mpatches.Patch(color="#DC2626", label="Negativa (↑ input → ↓ CUV)")
    ax.legend(handles=[pos_patch, neg_patch], fontsize=9)
    plt.tight_layout()
    return fig

def fig_scatter_inputs_vs_cuv(mc_res):
    si = mc_res["stochastic_inputs"]
    cuv = mc_res["outputs"]["CUV_t"]
    label_map = {"rho_r": ("Densidad roca ρr", "g/cm³"), "rho_e": ("Densidad explosivo ρe", "g/cm³"),
                 "e_expl": ("Energía específica", "MJ/kg"), "Dh": ("Diámetro taladro", "mm"),
                 "Ce": ("Costo explosivo", "USD/kg"), "C_init": ("Costo iniciador", "USD/u"),
                 "Cp": ("Costo perforación", "USD/m")}
    keys = list(si.keys())
    ncols = 3; nrows = math.ceil(len(keys) / ncols)
    fig, axes = plt.subplots(nrows, ncols, figsize=(14, 4.5 * nrows))
    fig.suptitle("Scatter: Variables de Entrada vs CUV [USD/t]", fontsize=11, fontweight="bold")
    axes_flat = axes.flatten()
    idx_sample = np.random.choice(len(cuv), size=min(2000, len(cuv)), replace=False)
    for i, key in enumerate(keys):
        ax = axes_flat[i]
        lbl, unit = label_map.get(key, (key, ""))
        ax.scatter(si[key][idx_sample], cuv[idx_sample], s=5, alpha=0.25, color="#2563EB")
        r, _ = stats.pearsonr(si[key], cuv)
        ax.set_xlabel(f"{lbl} [{unit}]", fontsize=9)
        ax.set_ylabel("CUV [USD/t]", fontsize=9)
        ax.set_title(f"r = {r:.3f}", fontsize=9)
    for j in range(len(keys), len(axes_flat)):
        axes_flat[j].set_visible(False)
    plt.tight_layout()
    return fig

def fig_burden_vs_diameter(p, det_res):
    Dh_range = np.linspace(50, 400, 200)
    burdens = 0.012 * math.sqrt(2.0 * p["rho_e"] / p["rho_r"]) * Dh_range
    spacings = p["Ks"] * burdens
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.suptitle("Burden y Espaciamiento vs Diámetro del Taladro", fontsize=11, fontweight="bold")
    ax.plot(Dh_range, burdens, color="#2563EB", lw=2.5, label="Burden B [m]")
    ax.plot(Dh_range, spacings, color="#10B981", lw=2.5, ls="--", label=f"S = {p['Ks']}·B [m]")
    ax.axvline(p["Dh"], color="#DC2626", lw=1.5, ls=":", label=f"Dh actual = {p['Dh']} mm")
    ax.axhline(det_res["geom"]["B"], color="#F59E0B", lw=1.5, ls=":", label=f"B actual = {det_res['geom']['B']:.2f} m")
    ax.scatter([p["Dh"]], [det_res["geom"]["B"]], s=80, color="#DC2626", zorder=5)
    ax.set_xlabel("Diámetro del taladro Dh [mm]"); ax.set_ylabel("Distancia [m]")
    ax.legend(); plt.tight_layout()
    return fig

def fig_cost_breakdown(p, det_res):
    Ce_item = det_res["Qe"] * p["Ce"]
    CCD_item = det_res["geom"]["Lc"] * p["C_CD"]
    Cinit_item = p["N_init"] * p["C_init"]
    Cperf_item = det_res["costs"]["Cp_t"]
    labels = ["Explosivo principal", "Cordón detonante", "Iniciadores", "Perforación"]
    values = [Ce_item, CCD_item, Cinit_item, Cperf_item]
    colors = ["#2563EB", "#0EA5E9", "#7C3AED", "#10B981"]
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    fig.suptitle("Desglose de Costos por Taladro", fontsize=11, fontweight="bold")
    axes[0].pie(values, labels=labels, colors=colors, autopct="%1.1f%%", startangle=90,
                wedgeprops={"edgecolor": "white", "linewidth": 2})
    axes[0].set_title("Distribución porcentual")
    axes[1].bar(labels, values, color=colors, edgecolor="white")
    for i, v in enumerate(values):
        axes[1].text(i, v + 0.5, f"${v:.1f}", ha="center", va="bottom", fontsize=9)
    axes[1].set_ylabel("Costo [USD/taladro]"); axes[1].set_title("Costo absoluto por ítem")
    plt.xticks(rotation=15); plt.tight_layout()
    return fig

def fig_fc_sensitivity(p, det_res):
    Dh_range = np.linspace(80, 350, 150)
    rho_r_vals = [1.9, 2.3, 2.7, 3.2]
    colors_r = ["#60A5FA", "#3B82F6", "#1D4ED8", "#1E3A8A"]
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.suptitle("Factor de Carga FC [kg/t] vs Diámetro del Taladro", fontsize=11, fontweight="bold")
    for rr, col in zip(rho_r_vals, colors_r):
        fc_list = []
        for dh in Dh_range:
            pp = dict(p); pp["Dh"] = dh; pp["rho_r"] = rr
            g = calc_geometry(pp); qe = calc_explosive_mass(pp, g)
            lf = calc_load_factor(pp, g, qe)
            fc_list.append(lf["FC"])
        ax.plot(Dh_range, fc_list, color=col, lw=2, label=f"ρr = {rr} g/cm³")
    ax.axvline(p["Dh"], color="#DC2626", lw=1.5, ls="--", label=f"Dh actual = {p['Dh']} mm")
    ax.axhline(det_res["lf"]["FC"], color="#F59E0B", lw=1.5, ls=":", label=f"FC actual = {det_res['lf']['FC']:.3f} kg/t")
    ax.set_xlabel("Diámetro del taladro Dh [mm]"); ax.set_ylabel("Factor de carga FC [kg/t]")
    ax.legend(); plt.tight_layout()
    return fig

def fig_cuv_sensitivity_cp_ce(p):
    cp_range = np.linspace(5, 50, 60)
    ce_range = np.linspace(0.3, 3.0, 60)
    CP, CE = np.meshgrid(cp_range, ce_range)
    CUV_grid = np.zeros_like(CP)
    for i in range(len(ce_range)):
        for j in range(len(cp_range)):
            pp = dict(p); pp["Cp"] = float(cp_range[j]); pp["Ce"] = float(ce_range[i])
            r = calc_all(pp)
            CUV_grid[i, j] = r["bank"]["CUV_t"]
    fig, ax = plt.subplots(figsize=(9, 6))
    fig.suptitle("Mapa de Contorno: CUV [USD/t] vs Cp y Ce", fontsize=11, fontweight="bold")
    cs = ax.contourf(CP, CE, CUV_grid, levels=20, cmap="RdYlGn_r")
    plt.colorbar(cs, ax=ax, label="CUV [USD/t]")
    ax.contour(CP, CE, CUV_grid, levels=10, colors="white", linewidths=0.5, alpha=0.5)
    ax.scatter([p["Cp"]], [p["Ce"]], s=120, color="white", marker="*", zorder=5,
               label=f"Caso actual: Cp={p['Cp']}, Ce={p['Ce']}")
    ax.set_xlabel("Costo perforación Cp [USD/m]"); ax.set_ylabel("Costo explosivo Ce [USD/kg]")
    ax.legend(fontsize=9); plt.tight_layout()
    return fig

def fig_mc_stats_comparison(mc_res, det_res):
    keys_to_plot = ["B", "Qe", "FC", "CT", "CUV_t"]
    labels_map = {"B": "B [m]", "Qe": "Qe [kg]", "FC": "FC [kg/t]",
                  "CT": "CT [USD]", "CUV_t": "CUV [USD/t]"}
    det_vals = {"B": det_res["geom"]["B"], "Qe": det_res["Qe"],
                "FC": det_res["lf"]["FC"], "CT": det_res["bank"]["CT"],
                "CUV_t": det_res["bank"]["CUV_t"]}
    fig, axes = plt.subplots(1, len(keys_to_plot), figsize=(15, 5))
    fig.suptitle("Comparación Determinístico vs Monte Carlo (P5–P95)", fontsize=11, fontweight="bold")
    for i, key in enumerate(keys_to_plot):
        ax = axes[i]
        mc_st = mc_res["mc_stats"][key]
        yvals = [mc_st["p5"], mc_st["p50"], mc_st["mean"], mc_st["p90"], mc_st["p95"]]
        xlabels = ["P5", "P50", "Media", "P90", "P95"]
        colors_b = ["#60A5FA", "#3B82F6", "#10B981", "#F59E0B", "#EF4444"]
        bars = ax.bar(xlabels, yvals, color=colors_b, edgecolor="white")
        ax.axhline(det_vals[key], color="#DC2626", lw=2, ls="--", label=f"Det.={det_vals[key]:.3f}")
        ax.set_title(labels_map[key], fontsize=9); ax.legend(fontsize=7)
        for bar, val in zip(bars, yvals):
            ax.text(bar.get_x() + bar.get_width() / 2, val, f"{val:.2f}",
                    ha="center", va="bottom", fontsize=6.5)
    plt.tight_layout()
    return fig

def fig_explosive_comparison(p):
    explosives = {
        "ANFO seco": {"rho_e": 0.80, "e": 3.85},
        "ANFO pesado 70/30": {"rho_e": 1.10, "e": 3.50},
        "Emulsión bombeada": {"rho_e": 1.20, "e": 3.20},
        "Emulsión encartuchada": {"rho_e": 1.25, "e": 3.30},
        "ANFO+Emulsión 50/50": {"rho_e": 1.00, "e": 3.60},
    }
    e_ANFO, rho_ANFO = 3.85, 0.80
    names, rws_vals, rbs_vals, burden_vals, fc_vals = [], [], [], [], []
    for name, props in explosives.items():
        rws = (props["e"] / e_ANFO) * 100
        rbs = rws * (props["rho_e"] / rho_ANFO)
        pp = dict(p); pp["rho_e"] = props["rho_e"]
        g = calc_geometry(pp); qe = calc_explosive_mass(pp, g); lf = calc_load_factor(pp, g, qe)
        names.append(name); rws_vals.append(rws); rbs_vals.append(rbs)
        burden_vals.append(g["B"]); fc_vals.append(lf["FC"])
    x = np.arange(len(names))
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    fig.suptitle("Comparación de Explosivos Comerciales", fontsize=11, fontweight="bold")
    axes[0, 0].bar(x, rws_vals, color="#2563EB", edgecolor="white")
    axes[0, 0].axhline(100, color="#DC2626", ls="--", lw=1.5, label="ANFO=100%")
    axes[0, 0].set_xticks(x); axes[0, 0].set_xticklabels(names, fontsize=7, rotation=10)
    axes[0, 0].set_title("RWS [%]"); axes[0, 0].legend()
    axes[0, 1].bar(x, rbs_vals, color="#10B981", edgecolor="white")
    axes[0, 1].axhline(100, color="#DC2626", ls="--", lw=1.5, label="ANFO=100%")
    axes[0, 1].set_xticks(x); axes[0, 1].set_xticklabels(names, fontsize=7, rotation=10)
    axes[0, 1].set_title("RBS [%]"); axes[0, 1].legend()
    axes[1, 0].bar(x, burden_vals, color="#7C3AED", edgecolor="white")
    axes[1, 0].bar(x[1], burden_vals[1], color="#F59E0B", edgecolor="white", label="Actual")
    axes[1, 0].set_xticks(x); axes[1, 0].set_xticklabels(names, fontsize=7, rotation=10)
    axes[1, 0].set_title("Burden B [m]"); axes[1, 0].legend()
    axes[1, 1].bar(x, fc_vals, color="#EF4444", edgecolor="white")
    axes[1, 1].set_xticks(x); axes[1, 1].set_xticklabels(names, fontsize=7, rotation=10)
    axes[1, 1].set_title("Factor de Carga FC [kg/t]")
    plt.tight_layout()
    return fig

# ─────────────────────────────────────────────
# EXPORTAR EXCEL (en memoria)
# ─────────────────────────────────────────────
def export_excel_bytes(p, det_res, mc_res):
    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    sub_fill = PatternFill("solid", fgColor="BDD7EE")
    sub_font = Font(bold=True, color="000000", size=10)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FFCC99")
    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def hdr_row(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

    def sub_row(ws, row, values):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = sub_fill; c.font = sub_font; c.alignment = center; c.border = border

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja 1 — Datos de Entrada
    ws1 = wb.active; ws1.title = "1_Datos_Entrada"
    hdr_row(ws1, 1, ["Parámetro / Símbolo", "Valor", "Unidades", "Valores típicos", "CoV MC (%)"])
    sections = [
        ("SECCIÓN 1 — ROCA Y GEOMECÁNICA", ["rho_r"]),
        ("SECCIÓN 2 — GEOMETRÍA DEL BANCO", ["H", "alpha", "Ks", "Kj", "Kt"]),
        ("SECCIÓN 3 — TALADRO Y EXPLOSIVO", ["Dh", "rho_e", "e_expl"]),
        ("SECCIÓN 4 — COSTOS UNITARIOS", ["Ce", "C_init", "N_init", "C_CD", "Cp"]),
        ("SECCIÓN 5 — PARÁMETROS OPERACIONALES", ["Nt", "Fop"]),
        ("SECCIÓN 6 — SIMULACIÓN MC", ["n_sim"]),
    ]
    row = 2
    for sec_title, keys in sections:
        sub_row(ws1, row, [sec_title, "", "", "", ""])
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1
        for k in keys:
            info = PARAM_INFO[k]
            cov = int(info["std_frac"] * 100) if info["stochastic"] else "—"
            for col, v in enumerate([f"{info['label']} [{info['symbol']}]",
                                     p[k], info["unit"], info["hint"], cov], 1):
                c = ws1.cell(row=row, column=col, value=v)
                c.alignment = left; c.border = border
            row += 1
    set_col_widths(ws1, [36, 12, 12, 55, 10])

    # Hoja 2 — Resultados
    ws2 = wb.create_sheet("2_Resultados")
    hdr_row(ws2, 1, ["Resultado", "Valor", "Unidades", "Interpretación"])
    results = [
        ("Burden B", f"{det_res['geom']['B']:.3f}", "m", interpret_burden(det_res['geom']['B'])[1]),
        ("Espaciamiento S", f"{det_res['geom']['S']:.3f}", "m", ""),
        ("Sobreperforación J", f"{det_res['geom']['J']:.3f}", "m", ""),
        ("Longitud taco Ltaco", f"{det_res['geom']['Ltaco']:.3f}", "m", ""),
        ("Longitud taladro Lt", f"{det_res['geom']['Lt']:.3f}", "m", ""),
        ("Longitud de carga Lc", f"{det_res['geom']['Lc']:.3f}", "m", ""),
        ("Masa explosivo Qe", f"{det_res['Qe']:.2f}", "kg/taladro", ""),
        ("Factor de carga FC", f"{det_res['lf']['FC']:.4f}", "kg/t", interpret_fc(det_res['lf']['FC'])[2]),
        ("RWS", f"{det_res['energy']['RWS']:.1f}", "%", ""),
        ("RBS", f"{det_res['energy']['RBS']:.1f}", "%", interpret_rbs(det_res['energy']['RBS'])[1]),
        ("Costo explosivo+acces./taladro", f"{det_res['costs']['Ce_t']:.2f}", "USD", ""),
        ("Costo perforación/taladro", f"{det_res['costs']['Cp_t']:.2f}", "USD", ""),
        ("Costo total banco CT", f"{det_res['bank']['CT']:,.0f}", "USD", ""),
        ("Producción total Ttotal", f"{det_res['bank']['Ttotal']:,.0f}", "t", ""),
        ("CUV (por tonelada)", f"{det_res['bank']['CUV_t']:.4f}", "USD/t", interpret_cuv(det_res['bank']['CUV_t'])[2]),
        ("CUV (por m³)", f"{det_res['bank']['CUV_vol']:.4f}", "USD/m³", ""),
    ]
    for r_i, row_data in enumerate(results, 2):
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=r_i, column=col, value=val)
            c.alignment = left; c.border = border
    set_col_widths(ws2, [32, 14, 12, 55])

    # Hoja 3 — Monte Carlo
    ws3 = wb.create_sheet("3_MonteCarlo")
    hdr_row(ws3, 1, ["Variable", "Unidades", "Media", "Desv. Est.", "P5", "P50", "P90", "P95", "CV (%)"])
    out_labels = {"B": ("Burden B", "m"), "Qe": ("Masa explosivo", "kg"),
                  "FC": ("Factor carga FC", "kg/t"), "RWS": ("RWS", "%"), "RBS": ("RBS", "%"),
                  "CT": ("Costo total CT", "USD"), "CUV_t": ("CUV por tonelada", "USD/t"),
                  "CUV_vol": ("CUV por m³", "USD/m³"), "Ttotal": ("Producción total", "t")}
    r3 = 2
    for key, (lbl, unit) in out_labels.items():
        mc_st = mc_res["mc_stats"].get(key, {})
        if not mc_st: continue
        cv = mc_st["std"] / mc_st["mean"] * 100 if mc_st["mean"] != 0 else 0
        for col, v in enumerate([lbl, unit, mc_st["mean"], mc_st["std"], mc_st["p5"],
                                  mc_st["p50"], mc_st["p90"], mc_st["p95"], cv], 1):
            c = ws3.cell(row=r3, column=col, value=v)
            c.alignment = left; c.border = border
            if col >= 3: c.number_format = "0.0000"
        r3 += 1
    set_col_widths(ws3, [26, 12, 14, 14, 12, 12, 12, 12, 10])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────
# APP STREAMLIT
# ─────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Análisis de Voladura",
        page_icon="💥",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.title("💥 Análisis de Costos y Consumo de Explosivos por Voladura")
    st.markdown(
        "**Método:** Langefors–Kihlström (1963) + Ash (1963)  ·  "
        "**Simulación:** Monte Carlo integrada"
    )
    st.divider()

    # ── SIDEBAR ──────────────────────────────
    with st.sidebar:
        st.header("⚙️ Parámetros de Entrada")

        st.subheader("🪨 Roca y Geomecánica")
        rho_r = st.number_input("Densidad de la roca ρr [g/cm³]",
                                 min_value=1.5, max_value=4.5, value=2.70, step=0.05,
                                 help="Muy blanda 1.8–2.2 · Media 2.2–2.5 · Dura 2.5–2.9 · Ultra-dura 3.0–4.0")

        st.subheader("📐 Geometría del Banco")
        H      = st.number_input("Altura del banco H [m]", 1.0, 30.0, 10.0, 0.5,
                                  help="Banco bajo 5–10 · Estándar 10–15 · Alto 15–20")
        alpha  = st.number_input("Ángulo del taladro α [°]", 55.0, 90.0, 80.0, 1.0,
                                  help="Vertical 90° · Inclinación leve 75–85° · Moderada 65–75°")
        Ks     = st.number_input("Coef. espaciamiento Ks", 0.9, 1.8, 1.15, 0.01,
                                  help="Roca fracturada 1.0–1.1 · Estándar 1.15 · Competente 1.2–1.5")
        Kj     = st.number_input("Coef. sobreperforación Kj", 0.15, 0.50, 0.30, 0.01,
                                  help="Roca blanda 0.20–0.25 · Estándar 0.30 · Roca dura 0.35–0.40")
        Kt     = st.number_input("Coef. taco (stemming) Kt", 0.60, 1.10, 0.75, 0.01,
                                  help="Máx. eficiencia 0.70–0.75 · Estándar 0.75–0.80")

        st.subheader("🔩 Taladro y Explosivo")
        Dh     = st.number_input("Diámetro del taladro Dh [mm]", 50.0, 450.0, 165.0, 5.0,
                                  help="Jack-leg 89–115 · Jumbo mediano 115–165 · Rotary grande 251–381")
        rho_e  = st.number_input("Densidad del explosivo ρe [g/cm³]", 0.70, 1.50, 1.10, 0.05,
                                  help="ANFO seco 0.80 · ANFO pesado 1.00–1.10 · Emulsión 1.15–1.25")
        e_expl = st.number_input("Energía específica e_expl [MJ/kg]", 2.0, 5.0, 3.50, 0.05,
                                  help="ANFO seco 3.85 · ANFO pesado 3.50 · Emulsión 3.20–3.30")

        st.subheader("💰 Costos Unitarios")
        Ce     = st.number_input("Costo explosivo Ce [USD/kg]", 0.20, 3.50, 0.95, 0.05,
                                  help="ANFO seco 0.45–0.70 · ANFO pesado 0.80–1.20 · Emulsión 0.90–2.50")
        C_init = st.number_input("Costo iniciador C_init [USD/u]", 0.50, 30.0, 4.50, 0.25,
                                  help="Fulminante 0.50–1.50 · Nonel 2.00–8.00 · Electrónico 8.00–25.00")
        N_init = st.number_input("Iniciadores por taladro N_init", 1, 3, 1, 1,
                                  help="Simple 1 · Doble >15 m: 2 · Triple roca muy dura: 2–3")
        C_CD   = st.number_input("Costo cordón detonante C_CD [USD/m]", 0.0, 0.60, 0.20, 0.01,
                                  help="Ligero 0.10–0.18 · Estándar 0.18–0.28 · Sin cordón 0.00")
        Cp     = st.number_input("Costo perforación Cp [USD/m]", 4.0, 60.0, 18.0, 0.5,
                                  help="Jack-leg 6–15 · Jumbo mediano 12–22 · Rotary mediano 18–32")

        st.subheader("🏭 Parámetros Operacionales")
        Nt     = st.number_input("N° taladros por banco Nt", 5, 800, 48, 1,
                                  help="Pequeño 10–30 · Mediano 30–80 · Grande 80–200")
        Fop    = st.number_input("Overhead operacional Fop [%]", 0.0, 30.0, 12.0, 0.5,
                                  help="Mínimo 8–10 · Estándar 10–15 · Alto 15–20")

        st.subheader("🎲 Monte Carlo")
        n_sim  = st.selectbox("Número de simulaciones", [5000, 10000, 50000], index=1,
                               help="Rápido 5,000 · Estándar 10,000 · Alta precisión 50,000")

        run_btn = st.button("▶ Calcular", type="primary", use_container_width=True)

    # ── CÁLCULO ──────────────────────────────
    if not run_btn:
        st.info("👈 Ingresa los parámetros en el panel izquierdo y presiona **Calcular**.")
        return

    p = {
        "rho_r": float(rho_r), "rho_e": float(rho_e), "e_expl": float(e_expl),
        "Dh": float(Dh), "H": float(H), "alpha": float(alpha),
        "Ks": float(Ks), "Kj": float(Kj), "Kt": float(Kt),
        "Ce": float(Ce), "C_init": float(C_init), "N_init": int(N_init),
        "C_CD": float(C_CD), "Cp": float(Cp), "Nt": int(Nt),
        "Fop": float(Fop), "n_sim": int(n_sim),
    }

    with st.spinner("⏳ Ejecutando cálculos..."):
        det_res = calc_all(p)
        mc_res  = monte_carlo(p)

    geom   = det_res["geom"]
    energy = det_res["energy"]
    lf     = det_res["lf"]
    costs  = det_res["costs"]
    bank   = det_res["bank"]

    cuv_cat, _, cuv_msg = interpret_cuv(bank["CUV_t"])
    fc_cat,  _, fc_msg  = interpret_fc(lf["FC"])
    color_map = {"verde": "normal", "amarillo": "off", "naranja": "inverse", "rojo": "inverse"}

    # ── MÉTRICAS PRINCIPALES ─────────────────
    st.subheader("📊 Resultados Determinísticos")
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Burden B", f"{geom['B']:.3f} m")
    c2.metric("Espaciamiento S", f"{geom['S']:.3f} m")
    c3.metric("Masa explosivo Qe", f"{det_res['Qe']:.2f} kg")
    c4.metric("Factor de carga FC", f"{lf['FC']:.4f} kg/t")
    c5.metric("CUV", f"{bank['CUV_t']:.4f} USD/t")

    c6, c7, c8, c9, c10 = st.columns(5)
    c6.metric("Long. taladro Lt", f"{geom['Lt']:.2f} m")
    c7.metric("Long. carga Lc", f"{geom['Lc']:.2f} m")
    c8.metric("RWS", f"{energy['RWS']:.1f} %")
    c9.metric("RBS", f"{energy['RBS']:.1f} %")
    c10.metric("Costo total CT", f"${bank['CT']:,.0f} USD")

    # Interpretaciones
    st.markdown(f"**CUV → {cuv_cat}:** {cuv_msg}  |  **FC → {fc_cat}:** {fc_msg}")

    # ── TABS ─────────────────────────────────
    tabs = st.tabs([
        "📈 Histogramas MC",
        "🌪️ Tornado & Scatter",
        "🔍 Sensibilidad",
        "💣 Comparación Explosivos",
        "📊 Stats MC",
        "📥 Exportar Excel",
    ])

    with tabs[0]:
        st.subheader("Distribuciones Monte Carlo")
        col_a, col_b = st.columns(2)
        with col_a:
            st.pyplot(fig_histogram_cdf(mc_res["outputs"]["CUV_t"], "CUV", "USD/t", bank["CUV_t"]))
            st.pyplot(fig_histogram_cdf(mc_res["outputs"]["CT"], "CT", "USD", bank["CT"]))
        with col_b:
            st.pyplot(fig_histogram_cdf(mc_res["outputs"]["FC"], "FC", "kg/t", lf["FC"]))
            st.pyplot(fig_histogram_cdf(mc_res["outputs"]["B"], "B", "m", geom["B"]))
        st.pyplot(fig_mc_stats_comparison(mc_res, det_res))

    with tabs[1]:
        st.subheader("Diagrama de Tornado")
        st.pyplot(fig_tornado(mc_res["correlations"]))
        st.subheader("Scatter: Variables vs CUV")
        st.pyplot(fig_scatter_inputs_vs_cuv(mc_res))

    with tabs[2]:
        st.subheader("Burden y Espaciamiento vs Diámetro")
        st.pyplot(fig_burden_vs_diameter(p, det_res))
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            st.subheader("Factor de Carga FC vs Diámetro")
            st.pyplot(fig_fc_sensitivity(p, det_res))
        with col_s2:
            st.subheader("Desglose de Costos por Taladro")
            st.pyplot(fig_cost_breakdown(p, det_res))
        st.subheader("Mapa de Contorno CUV vs Cp & Ce")
        st.pyplot(fig_cuv_sensitivity_cp_ce(p))

    with tabs[3]:
        st.subheader("Comparación de Explosivos Comerciales")
        st.pyplot(fig_explosive_comparison(p))

    with tabs[4]:
        st.subheader("Estadísticas Monte Carlo Completas")
        out_labels = {
            "B": ("Burden B", "m"), "Qe": ("Masa explosivo Qe", "kg"),
            "FC": ("Factor de carga FC", "kg/t"), "RWS": ("RWS", "%"), "RBS": ("RBS", "%"),
            "CT": ("Costo total CT", "USD"), "CUV_t": ("CUV por tonelada", "USD/t"),
            "CUV_vol": ("CUV por m³", "USD/m³"), "Ttotal": ("Producción total", "t"),
        }
        rows = []
        for key, (lbl, unit) in out_labels.items():
            mc_st = mc_res["mc_stats"].get(key, {})
            if not mc_st: continue
            cv = mc_st["std"] / mc_st["mean"] * 100 if mc_st["mean"] != 0 else 0
            rows.append({
                "Variable": lbl, "Unidad": unit,
                "Media": round(mc_st["mean"], 4), "Desv. Est.": round(mc_st["std"], 4),
                "P5": round(mc_st["p5"], 4), "P50": round(mc_st["p50"], 4),
                "P90": round(mc_st["p90"], 4), "P95": round(mc_st["p95"], 4),
                "CV (%)": round(cv, 2),
            })
        import pandas as pd
        st_df = pd.DataFrame(rows)
        st.dataframe(st_df, use_container_width=True)

        st.subheader("Correlaciones de Pearson con CUV [USD/t]")
        corr_labels = {"rho_r": "Densidad roca ρr", "rho_e": "Densidad explosivo ρe",
                       "e_expl": "Energía específica e_expl", "Dh": "Diámetro taladro Dh",
                       "Ce": "Costo explosivo Ce", "C_init": "Costo iniciador C_init",
                       "Cp": "Costo perforación Cp"}
        corr_rows = []
        for k, r_val in sorted(mc_res["correlations"].items(), key=lambda x: abs(x[1]), reverse=True):
            interp = ("Positiva fuerte" if r_val > 0.5 else "Positiva moderada" if r_val > 0.2 else
                      "Positiva débil" if r_val > 0 else "Negativa débil" if r_val > -0.2 else
                      "Negativa moderada" if r_val > -0.5 else "Negativa fuerte")
            corr_rows.append({"Variable": corr_labels.get(k, k), "r": round(r_val, 4), "Interpretación": interp})
        st.dataframe(pd.DataFrame(corr_rows), use_container_width=True)

    with tabs[5]:
        st.subheader("📥 Descargar Reporte Excel")
        st.markdown("El archivo Excel incluye: **Datos de entrada**, **Resultados determinísticos**, "
                    "**Estadísticas Monte Carlo** y **Correlaciones**.")
        excel_bytes = export_excel_bytes(p, det_res, mc_res)
        st.download_button(
            label="⬇️ Descargar resultados_voladura.xlsx",
            data=excel_bytes,
            file_name="resultados_voladura.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

if __name__ == "__main__":
    main()
